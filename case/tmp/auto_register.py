__author__ = 'Administrator'
import random
from openpyxl import load_workbook
import requests


def output_num():
    """
    随机产生一个手机号码，返回手机号码类型为int
    """
    list = ["132","151","185","138"]
    header_num = list[random.randint(0,3)]
    taif_num = ''.join(random.sample(['0','1','2','3','4','5','6','7','8','9'],8))
    phonenum = header_num+taif_num
    phonenum = phonenum
    return phonenum

def write_data(filename,phonenum,row):
    """
    写回测试结果到excel中
    """
    wb = load_workbook(filename = filename)
    sheet = wb["register"]
    sheet.cell(row,1).value = phonenum
    sheet.cell(row,2).value = "123456"
    wb.save(filename = filename)
    wb.close()

def do_register(number,url,filename):
    """
    自动注册手机号码并将注册成功的手机号码写入excel中
    number：表示注册的手机号码个数，格式为int
    url:注册的url，格式为字符串
    filename:你想要写入的excel文件名，文件必须为xlsx结尾，文件名为字符串类型
    row：控制写入的行数
    """

    time = 0
    row=0
    while time < number:
        phonenumber = output_num()
        param = {"mobilephone":"","pwd":"tudou123456"}
        #密码默认为tudou123456
        param["mobilephone"] = phonenumber
        resp = requests.request("get",url=url,params=param)
        #注册方法默认为get
        if resp.json()["msg"] == "注册成功":
            print(resp.json())
            row = row +1
            time = time+1
            write_data(filename=filename,phonenum=str(phonenumber),row=row)



if __name__ == '__main__':
    number = 5
    url = "http://test.lemonban.com/futureloan/mvc/api/member/register"
    filename = "auto_register.xlsx"
    do_register(number=number,url=url,filename=filename)
